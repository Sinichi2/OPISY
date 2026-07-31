"""
    Extraction step: pull raw rows out of a staff-supplied Excel workbook.

    Nothing is cleaned here. Every sheet is read, every row comes back as a
    {header: value} dict of strings, exactly as it was typed.
"""
from openpyxl import load_workbook


def read_rows(path):
    """
        Yields one dict per data row, keyed by that sheet's header text.
        All sheets in the workbook are read.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            header = None
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c).strip() for c in row]
                if header is None:
                    # ponytail: header = first row with 2+ filled cells. Covers the
                    # title/logo rows real sheets have on top. If a sheet ever needs
                    # an explicit header row number, pass it in per sheet.
                    if sum(bool(c) for c in cells) >= 2:
                        header = cells
                    continue
                if any(cells):
                    yield dict(zip(header, cells))
    finally:
        wb.close()
